# core/management/commands/import_excel.py
"""
Import de donnees depuis un fichier Excel de suivi financier BTP.
Format attendu : SF-XXXX-XX-NomClient-Lieu.xlsx

Logique entre feuilles :
  - "Commande Materiau" = commandes passees et payees => stock entre (Achat)
  - "Etat stock"        = utilisation materiaux sur chantier => stock sort (BonSortie)
  - "DEPENSES"          = toutes les depenses (materiaux + main d'oeuvre + services)
  - "Versement"         = paiements du client

Classification stricte des intervenants :
  - Employe   : macon, ouvrier (Abdou Diouf) => table Employe uniquement
  - Sous-traitant : electricien, plombier, mouleur => table SousTraitant uniquement
  - Fournisseur : vendeur de materiaux => table Fournisseur uniquement
  - Frais : transport/carburant => Achat sans fournisseur
"""
import os
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from core.models import (
    Projet, TypeProjet, Proprietaire, Fournisseur,
    Materiel, CategorieMateriel, Achat, LigneAchat,
    Versement, PhaseVersement, Employe,
    SousTraitant, ContratSousTraitance,
    BonSortie, LigneBonSortie,
    CompteComptable, EcritureComptable, LigneEcriture,
    CompteBancaire, TransactionBancaire,
)
from core.services.comptabilite import (
    generer_ecriture_achat,
    generer_ecriture_versement,
    generer_ecriture_sous_traitance,
)


# ═══════════════════════════════════════════════════════════════════════════
# Utilitaires
# ═══════════════════════════════════════════════════════════════════════════

def to_decimal(val, default=Decimal("0")):
    if val is None:
        return default
    try:
        return Decimal(str(val).replace(" ", "").replace(",", "."))
    except (InvalidOperation, ValueError):
        return default


def to_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    try:
        return datetime.strptime(str(val).strip(), "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return None


def normalize(s):
    if not s:
        return ""
    return str(s).strip()


def fix_date(d, annee_projet=2025):
    """Corrige les dates placeholder (avant 2024) en gardant mois/jour."""
    if d and d.year < 2024:
        annee_corrigee = annee_projet + 1  # 2026
        try:
            return d.replace(year=annee_corrigee)
        except ValueError:
            return d.replace(year=annee_corrigee, day=28)
    return d


# ═══════════════════════════════════════════════════════════════════════════
# Mapping des fichiers vers les infos projet/client
# Format nom de fichier : SF-XXXX-XX-ClientKey-Lieu.xlsx
# ═══════════════════════════════════════════════════════════════════════════

PROJETS_MAP = {
    "adieng": {
        "prenom": "Aida", "nom": "Dieng", "sexe": "Femme",
        "lieu": "Thies", "projet": "Villa Aida - Thies",
    },
    "aaidara": {
        "prenom": "A.", "nom": "Aidara", "sexe": "Homme",
        "lieu": "SCAT Urbain", "projet": "Villa Aidara - SCAT Urbain",
    },
    "bngom": {
        "prenom": "Babou", "nom": "Ngom", "sexe": "Homme",
        "lieu": "Saly", "projet": "Villa NGOM - Saly",
    },
}


def parse_filename(basename):
    """
    Parse le nom de fichier SF-XXXX-XX-ClientKey-Lieu.xlsx
    Retourne les infos projet depuis PROJETS_MAP.
    """
    parts = basename.split("-")
    if len(parts) >= 4:
        client_key = parts[3].lower()
        if client_key in PROJETS_MAP:
            return PROJETS_MAP[client_key]
    return None


# ═══════════════════════════════════════════════════════════════════════════
# Classification des intervenants (3 fichiers)
# ═══════════════════════════════════════════════════════════════════════════
# Roles : "macon", "sous_traitant", "fournisseur", "frais"
# IMPORTANT : un macon n'est PAS un fournisseur, un sous-traitant non plus

INTERVENANTS_MAP = {
    # ── Employes / Macons ────────────────────────────────────────────────
    "abdou diouf": ("Abdou Diouf", "macon"),
    "mamadou mbaye": ("Mamadou Mbaye", "macon"),
    "mamadou baye": ("Mamadou Mbaye", "macon"),       # variation ecriture = même personne
    "abdoul aziz": ("Abdoul Aziz", "macon"),

    # ── Sous-traitants specialises ───────────────────────────────────────
    "sylla diop": ("Sylla Diop", "sous_traitant"),          # Electricien
    "sylla diop": ("Sylla Diop", "sous_traitant"),
    "aliou plombier": ("Aliou Plombier", "sous_traitant"),  # Plombier
    "modou mouleur": ("Modou Mouleur", "sous_traitant"),    # Mouleur
    "gora gueye": ("Gora Gueye", "sous_traitant"),          # Fabricant briques + sable

    # ── Fournisseurs ADieng ──────────────────────────────────────────────
    "ibou mbaye": ("Ibou Mbaye - Quincaillerie", "fournisseur"),
    "quincaillerie": ("Ibou Mbaye - Quincaillerie", "fournisseur"),
    "diassap briques suarl": ("DIASSAP BRIQUES SUARL", "fournisseur"),
    "mariama sow": ("Mariama Sow", "fournisseur"),
    "turque vendeur brique": ("Independance Investment SARL IIS", "fournisseur"),
    "el hadji badara sall": ("EL Hadji Badara Sall", "fournisseur"),
    "taiba": ("Taiba", "fournisseur"),
    "mor": ("Mor", "fournisseur"),
    "ndi\u00e8me smk building": ("Ndieme SMK Building", "fournisseur"),
    "ndi\u00e9m\u00e9 smk building": ("Ndieme SMK Building", "fournisseur"),
    "ndieme smk building": ("Ndieme SMK Building", "fournisseur"),

    # ── Fournisseurs AAidara ─────────────────────────────────────────────
    "gora sarr": ("Gora Sarr", "fournisseur"),              # Ciment + fer
    "abc industrie": ("ABC Industrie", "fournisseur"),       # Briques
    "omar b\u00e9ton mbour": ("Omar Beton Mbour", "fournisseur"),
    "omar beton mbour": ("Omar Beton Mbour", "fournisseur"),
    "ousmane sambou": ("Ousmane Sambou", "fournisseur"),
    "ousmane sambou - camberene": ("Ousmane Sambou", "fournisseur"),
    "sadio ndiaye": ("Sadio Ndiaye", "fournisseur"),         # Briques

    # ── Fournisseurs BNGOM ───────────────────────────────────────────────
    "abib leye": ("Abib Leye", "fournisseur"),
    "entreprise seye construction": ("Entreprise Seye Construction", "fournisseur"),
    "entreprise seye constrcution": ("Entreprise Seye Construction", "fournisseur"),
    "mbatt s\u00e8ne": ("Mbatt Sene", "fournisseur"),
    "mbatt sene": ("Mbatt Sene", "fournisseur"),
    "quincaillerie gaye et fr\u00e8re": ("Quincaillerie Gaye et Frere", "fournisseur"),
    "quincaillerie gaye et frere": ("Quincaillerie Gaye et Frere", "fournisseur"),
    "saer - gueye": ("Saer Gueye", "fournisseur"),
    "saer gueye": ("Saer Gueye", "fournisseur"),

    # ── Fournisseurs communs ─────────────────────────────────────────────
    "sen eau": ("Sen Eau", "fournisseur"),

    # ── Services / impression ────────────────────────────────────────────
    "imprimerie - binette": ("Imprimerie Binette", "fournisseur"),
    "binette": ("Imprimerie Binette", "fournisseur"),

    # ── Frais (pas une personne) ─────────────────────────────────────────
    "carburant et p\u00e9age": ("Frais de transport", "frais"),
    "carburant peage": ("Frais de transport", "frais"),
    "frais de transport": ("Frais de transport", "frais"),
    "frais de trans\u00e9rt": ("Frais de transport", "frais"),
    "frais de trans\u00f4rt": ("Frais de transport", "frais"),
    "mamadou ndiaye": ("Mamadou Ndiaye", "frais"),           # Gardiennage
    "mairie grand yoff": ("Mairie Grand Yoff", "frais"),     # Autorisation
    "versement": ("Versement", "frais"),                     # Erreur saisie
}

# Specialites sous-traitants
SPECIALITES_MAP = {
    "Sylla Diop": "electricite_cfa",
    "Aliou Plombier": "plomberie_sanitaire",
    "Modou Mouleur": "autre",       # Mouleur (fabrication briques/parpaings)
    "Gora Gueye": "gros_oeuvre",     # Fabrication briques + sable
}

# Description du fournisseur (specialite / produits)
FOURNISSEUR_DESC = {
    "Ibou Mbaye - Quincaillerie": {"moral": True, "entreprise": "Quincaillerie Ibou Mbaye", "specialite": "Quincaillerie (fer, ciment, materiel BTP)"},
    "DIASSAP BRIQUES SUARL": {"moral": True, "entreprise": "DIASSAP BRIQUES SUARL", "specialite": "Fabrication et vente de briques"},
    "Independance Investment SARL IIS": {"moral": True, "entreprise": "Independance Investment SARL IIS", "specialite": "Vente de briques"},
    "Ndieme SMK Building": {"moral": True, "entreprise": "Ndieme SMK Building", "specialite": "Fournisseur de gravier"},
    "Imprimerie Binette": {"moral": True, "entreprise": "Imprimerie Binette", "specialite": "Impression plans et documents"},
    "Sen Eau": {"moral": True, "entreprise": "Sen Eau", "specialite": "Fourniture d'eau"},
    "Mariama Sow": {"moral": False, "specialite": "Fournisseur de gravier"},
    "Mor": {"moral": False, "specialite": "Fournisseur de sable"},
    "Taiba": {"moral": False, "specialite": "Fournisseur de sable"},
    "EL Hadji Badara Sall": {"moral": False, "specialite": "Fournisseur materiel BTP"},
    # ── AAidara ──
    "Gora Sarr": {"moral": False, "specialite": "Fournisseur ciment et fer"},
    "ABC Industrie": {"moral": True, "entreprise": "ABC Industrie", "specialite": "Fabrication et vente de briques"},
    "Omar Beton Mbour": {"moral": True, "entreprise": "Omar Beton Mbour", "specialite": "Fournisseur de gravier basalte"},
    "Ousmane Sambou": {"moral": False, "specialite": "Fournisseur de sable et gravillon"},
    "Sadio Ndiaye": {"moral": False, "specialite": "Fournisseur de briques"},
    # ── BNGOM ──
    "Abib Leye": {"moral": False, "specialite": "Fournisseur materiaux BTP"},
    "Entreprise Seye Construction": {"moral": True, "entreprise": "Entreprise Seye Construction", "specialite": "Menuiserie metallique"},
    "Mbatt Sene": {"moral": False, "specialite": "Fournisseur materiaux BTP"},
    "Quincaillerie Gaye et Frere": {"moral": True, "entreprise": "Quincaillerie Gaye et Frere", "specialite": "Quincaillerie (fer, ciment, materiel BTP)"},
    "Saer Gueye": {"moral": False, "specialite": "Fournisseur materiaux BTP"},
}


def classify_intervenant(nom_raw):
    nom = normalize(nom_raw)
    if not nom:
        return ("", "inconnu")
    key = nom.lower()
    if key in INTERVENANTS_MAP:
        return INTERVENANTS_MAP[key]
    if any(kw in nom.upper() for kw in ["SUARL", "SARL", "SA ", "BUILDING"]):
        return (nom, "fournisseur")
    return (nom, "fournisseur")


# ═══════════════════════════════════════════════════════════════════════════
# Normalisation lots
# ═══════════════════════════════════════════════════════════════════════════

LOT_MAP = {
    "main d'oeuvre": "Main d'oeuvre",
    "main d'\u0153uvre": "Main d'oeuvre",
    "ciment": "Ciment", "ciment ": "Ciment",
    "fer": "Fer",
    "sable": "Sable", "sable remblai": "Sable",
    "gravier": "Gravier",
    "brique": "Brique",
    "eau": "Eau",
    "materiel": "Materiel", "mat\u00e9riel": "Materiel",
    "transport": "Transport",
    "impression": "Impression",
    "equipement": "Equipement", "\u00e9quipement": "Equipement",
    "plomberie": "Plomberie",
    "installation": "Electricite",
    "coffrage": "Coffrage",
    "d\u00e9broussaillage": "Debroussaillage",
    "fabrication brique": "Fabrication briques",
    "fabrication briques": "Fabrication briques",
    # ── Nouveaux lots AAidara / BNGOM ──
    "gardiennage": "Gardiennage",
    "protection mitoyen": "Protection mitoyen",
    "annexe": "Frais annexes",
    "autorisation": "Frais administratifs",
    "acha gilet": "Equipement securite",
    "menuiserie": "Menuiserie",
    "baricade": "Installation chantier",
    "materiel electricit\u00e9": "Electricite",
    "materiel electricite": "Electricite",
    "materiel plomberie": "Plomberie",
    "d\u00e9broussaillage": "Debroussaillage",
    "debroussaillage": "Debroussaillage",
}


def normalize_lot(lot):
    if not lot:
        return "Divers"
    lot = normalize(lot)
    return LOT_MAP.get(lot.lower().strip(), lot)


# ═══════════════════════════════════════════════════════════════════════════
# Normalisation materiaux (feuilles Commande/Etat stock)
# ═══════════════════════════════════════════════════════════════════════════

MATERIAU_MAP = {
    "sable": ("Sable", "Sable", "m3"),
    "sable ": ("Sable", "Sable", "m3"),
    "gravier": ("Gravier", "Gravier", "m3"),
    "ciment": ("Ciment", "Ciment", "Tonne"),
    "ciment ": ("Ciment", "Ciment", "Tonne"),
    "fer 10 - 12": ("Fer", "Fer 10-12", "kg"),
    "fer 10-12": ("Fer", "Fer 10-12", "kg"),
    "fer 12": ("Fer", "Fer 12", "kg"),
    "fer 10": ("Fer", "Fer 10", "kg"),
    "fer 8": ("Fer", "Fer 8", "kg"),
    "fer 6": ("Fer", "Fer 6", "kg"),
    "fer6": ("Fer", "Fer 6", "kg"),
    "fil de fer": ("Fer", "Fil de fer", "kg"),
}


def normalize_materiau_commande(designation):
    d = normalize(designation).lower()
    if d in MATERIAU_MAP:
        return MATERIAU_MAP[d]
    return ("Divers", designation, "forfait")


# ═══════════════════════════════════════════════════════════════════════════
# Extraction quantites depuis les designations
# ═══════════════════════════════════════════════════════════════════════════

def extract_quantity_and_details(designation, lot_norm):
    """Extraire quantite, unite, et details depuis la designation."""
    d = normalize(designation).lower()
    if not d:
        return 1, "forfait", ""

    # Patterns : "20 m3 sable", "1 tonne ciment", "200 kg fer 12", "15 sacs de ciment"
    # "1000" apres ":" pour briques, "4 fut", "3 camions laterite"
    patterns = [
        (r'(\d+(?:[.,]\d+)?)\s*(?:m3|m\u00b3)', 'm3'),
        (r'(\d+(?:[.,]\d+)?)\s*(?:tonne|tonnes|t\b)', 'Tonne'),
        (r'(\d+(?:[.,]\d+)?)\s*kg', 'kg'),
        (r'(\d+(?:[.,]\d+)?)\s*sacs?\b', 'sac'),
        (r'(\d+(?:[.,]\d+)?)\s*(?:fut|futs|f\u00fbt|f\u00fbts)', 'fut'),
        (r'(\d+(?:[.,]\d+)?)\s*camions?\b', 'camion'),
        (r'(\d+(?:[.,]\d+)?)\s*(?:rouleaux?)', 'rouleau'),
        (r'(\d+(?:[.,]\d+)?)\s*(?:botte)', 'botte'),
        (r':\s*(\d+)', 'unite'),         # ":1000" briques
    ]

    total_qty = 0
    unite_found = "forfait"
    details_parts = []

    # Chercher toutes les quantites dans la designation
    # Ex: "200 kg fer 12 - 100 kg fer 6 - 5 kilo fil de fer" => 305 kg
    for pattern, unite in patterns:
        for m in re.finditer(pattern, d):
            qty = float(m.group(1).replace(",", "."))
            if qty > 0:
                total_qty += qty
                unite_found = unite

    if total_qty > 0:
        return max(1, int(total_qty)), unite_found, designation

    return 1, "forfait", designation


# ═══════════════════════════════════════════════════════════════════════════
# Moyens de paiement
# ═══════════════════════════════════════════════════════════════════════════

def normalize_moyen_achat(moyen):
    if not moyen:
        return "especes"
    moyen = normalize(moyen).lower()
    if "wave" in moyen or "vawe" in moyen:
        return "autre"
    if "virement" in moyen:
        return "virement"
    if "ch\u00e8que" in moyen or "cheque" in moyen:
        return "cheque"
    return "especes"


def normalize_moyen_versement(moyen):
    if not moyen:
        return "especes"
    moyen = normalize(moyen).lower()
    if "wave" in moyen or "vawe" in moyen:
        return "wave"
    if "virement" in moyen:
        return "virement bancaire"
    if "ch\u00e8que" in moyen or "cheque" in moyen:
        return "cheque"
    return "especes"


# ═══════════════════════════════════════════════════════════════════════════
# Commande Django
# ═══════════════════════════════════════════════════════════════════════════

class Command(BaseCommand):
    help = "Importer les donnees depuis un fichier Excel de suivi financier BTP"

    def add_arguments(self, parser):
        parser.add_argument("fichier", type=str, help="Chemin du fichier Excel (.xlsx)")
        parser.add_argument("--nom-projet", type=str, default="")
        parser.add_argument("--client", type=str, default="")
        parser.add_argument("--localisation", type=str, default="")

    @transaction.atomic
    def handle(self, *args, **options):
        fichier = options["fichier"]
        if not os.path.exists(fichier):
            raise CommandError(f"Fichier introuvable : {fichier}")

        try:
            import openpyxl
        except ImportError:
            raise CommandError("openpyxl requis : pip install openpyxl")

        wb = openpyxl.load_workbook(fichier, data_only=True)
        basename = os.path.splitext(os.path.basename(fichier))[0]

        # ── Parsing dynamique du nom de fichier ─────────────────────────
        infos = parse_filename(basename)
        if not infos:
            raise CommandError(
                f"Nom de fichier non reconnu : {basename}\n"
                f"Format attendu : SF-XXXX-XX-ClientKey-Lieu.xlsx\n"
                f"Clients connus : {', '.join(PROJETS_MAP.keys())}"
            )

        prenom_client = options["client"].split()[0] if options["client"] else infos["prenom"]
        nom_client = options["client"].split()[-1] if options["client"] else infos["nom"]
        nom_client_display = f"{prenom_client} {nom_client}"
        nom_projet = options["nom_projet"] or infos["projet"]
        localisation = options["localisation"] or infos["lieu"]
        sexe_client = infos.get("sexe", "Homme")

        self.stdout.write(f"\n{'='*60}")
        self.stdout.write(f"Import : {basename}")
        self.stdout.write(f"Projet : {nom_projet}")
        self.stdout.write(f"Client : {nom_client_display}")
        self.stdout.write(f"Lieu   : {localisation}")
        self.stdout.write(f"{'='*60}\n")

        # ── 1. Type de projet ────────────────────────────────────────────
        type_proj, _ = TypeProjet.objects.get_or_create(
            nom="Villa R+1",
            defaults={"description": "Villa residentielle R+1"}
        )

        # ── 2. Proprietaire ─────────────────────────────────────────────
        client, created = Proprietaire.objects.get_or_create(
            nom=nom_client,
            prenom=prenom_client,
            defaults={
                "telephone": "",
                "adresse": f"{localisation}, Senegal",
                "sexe": sexe_client,
                "est_moral": False,
            }
        )
        if created:
            self.stdout.write(self.style.SUCCESS(
                f"  + Proprietaire : {nom_client_display} ({sexe_client})"
            ))

        # ── 3. Projet ───────────────────────────────────────────────────
        projet, created = Projet.objects.get_or_create(
            nom=nom_projet,
            defaults={
                "type_projet": type_proj,
                "proprietaire": client,
                "localisation": localisation,
                "statut": "En cours",
                "date_debut": date(2025, 9, 1),
            }
        )
        if created:
            self.stdout.write(self.style.SUCCESS(f"  + Projet : {nom_projet}"))

        # ── 4. Categories et materiaux ──────────────────────────────────
        categories_cache = {}
        materiaux_cache = {}

        lots_materiaux = {
            "Ciment": ("Ciment", "Tonne"),
            "Fer": ("Fer a beton", "kg"),
            "Sable": ("Sable", "m3"),
            "Gravier": ("Gravier", "m3"),
            "Brique": ("Brique", "unite"),
            "Eau": ("Eau", "forfait"),
            "Coffrage": ("Bois de coffrage", "forfait"),
            "Plomberie": ("Materiel plomberie", "forfait"),
            "Electricite": ("Materiel electrique", "forfait"),
            "Equipement": ("Equipement chantier", "forfait"),
            "Materiel": ("Materiel divers", "forfait"),
            "Transport": ("Transport", "forfait"),
            "Impression": ("Impression / Plans", "forfait"),
            "Debroussaillage": ("Debroussaillage", "forfait"),
            "Fabrication briques": ("Fabrication briques", "unite"),
            "Main d'oeuvre": ("Main d'oeuvre", "forfait"),
            "Divers": ("Divers", "forfait"),
        }

        for lot_nom, (mat_nom, unite) in lots_materiaux.items():
            cat, _ = CategorieMateriel.objects.get_or_create(nom=lot_nom)
            categories_cache[lot_nom] = cat
            mat, _ = Materiel.objects.get_or_create(
                nom=mat_nom, defaults={"unite": unite, "categorie": cat}
            )
            materiaux_cache[lot_nom] = mat

        def get_or_create_materiel(cat_nom, mat_nom, unite="forfait"):
            cat = categories_cache.get(cat_nom)
            if not cat:
                cat, _ = CategorieMateriel.objects.get_or_create(nom=cat_nom)
                categories_cache[cat_nom] = cat
            key = f"{cat_nom}:{mat_nom}"
            mat = materiaux_cache.get(key)
            if not mat:
                mat, _ = Materiel.objects.get_or_create(
                    nom=mat_nom, defaults={"unite": unite, "categorie": cat}
                )
                materiaux_cache[key] = mat
            return mat

        # ── 5. Caches intervenants ──────────────────────────────────────
        fournisseurs_cache = {}
        employes_cache = {}
        sous_traitants_cache = {}
        contrats_st_data = {}

        def get_or_create_employe(nom_affiche):
            if nom_affiche in employes_cache:
                return employes_cache[nom_affiche]
            nom_parts = nom_affiche.split(" ", 1)
            prenom = nom_parts[0] if len(nom_parts) == 2 else ""
            nom_family = nom_parts[1] if len(nom_parts) == 2 else nom_affiche
            emp = Employe.objects.filter(nom=nom_family, prenom=prenom).first()
            if not emp:
                emp = Employe(
                    nom=nom_family, prenom=prenom,
                    poste="Macon", date_embauche=date(2025, 9, 1),
                    sexe="Homme", adresse=localisation, actif=True,
                )
                emp.save()
                self.stdout.write(f"  + Employe (Macon) : {nom_affiche}")
            employes_cache[nom_affiche] = emp
            return emp

        def get_or_create_sous_traitant(nom_affiche):
            if nom_affiche in sous_traitants_cache:
                return sous_traitants_cache[nom_affiche]
            st = SousTraitant.objects.filter(nom=nom_affiche).first()
            if not st:
                specialite = SPECIALITES_MAP.get(nom_affiche, "autre")
                st = SousTraitant.objects.create(
                    nom=nom_affiche, specialite=specialite,
                    adresse=localisation, actif=True,
                )
                self.stdout.write(f"  + Sous-traitant ({st.get_specialite_display()}) : {nom_affiche}")
            sous_traitants_cache[nom_affiche] = st
            return st

        def get_or_create_fournisseur(nom_affiche):
            if nom_affiche in fournisseurs_cache:
                return fournisseurs_cache[nom_affiche]

            desc = FOURNISSEUR_DESC.get(nom_affiche, {})
            est_moral = desc.get("moral", False)

            # Chercher existant
            if est_moral:
                entreprise_nom = desc.get("entreprise", nom_affiche)
                four = Fournisseur.objects.filter(entreprise=entreprise_nom).first()
            else:
                nom_parts = nom_affiche.split(" ", 1)
                prenom = nom_parts[0] if len(nom_parts) == 2 else ""
                nom_family = nom_parts[1] if len(nom_parts) == 2 else nom_affiche
                four = Fournisseur.objects.filter(nom=nom_family, prenom=prenom).first()

            if not four:
                if est_moral:
                    entreprise_nom = desc.get("entreprise", nom_affiche)
                    four = Fournisseur.objects.create(
                        est_moral=True, entreprise=entreprise_nom,
                        nom="", numero_identite=None, adresse=localisation,
                    )
                else:
                    nom_parts = nom_affiche.split(" ", 1)
                    prenom = nom_parts[0] if len(nom_parts) == 2 else ""
                    nom_family = nom_parts[1] if len(nom_parts) == 2 else nom_affiche
                    four = Fournisseur.objects.create(
                        est_moral=False, nom=nom_family, prenom=prenom,
                        numero_identite=None, adresse=localisation,
                    )
                spec = desc.get("specialite", "")
                label = f"  + Fournisseur : {nom_affiche}"
                if spec:
                    label += f" ({spec})"
                self.stdout.write(label)

            fournisseurs_cache[nom_affiche] = four
            return four

        # ── 6. VERSEMENTS ───────────────────────────────────────────────
        nb_versements = 0
        derniere_date_v = projet.date_debut  # fallback pour dates manquantes

        if "Versement" in wb.sheetnames:
            ws = wb["Versement"]
            for row in ws.iter_rows(min_row=4, max_row=ws.max_row, max_col=11, values_only=True):
                no, etape, designation, moyen, date_val, montant, *rest = row
                montant_d = to_decimal(montant)
                if montant_d <= 0:
                    continue

                date_v = to_date(date_val)
                date_v = fix_date(date_v)  # corrige dates < 2024 → 2026

                if not date_v:
                    date_v = derniere_date_v
                    self.stdout.write(self.style.WARNING(
                        f"  ! Versement {montant_d:,.0f} sans date => fallback {date_v}"
                    ))
                else:
                    derniere_date_v = date_v

                etape_nom = normalize(etape) or "Versement"
                designation_txt = normalize(designation) or f"Versement no{no}"

                # Convertir no en entier (peut etre "1-1", "2-3", etc.)
                try:
                    ordre_val = int(str(no).split("-")[0]) if no else nb_versements + 1
                except (ValueError, TypeError):
                    ordre_val = nb_versements + 1

                phase, _ = PhaseVersement.objects.get_or_create(
                    projet=projet, libelle=etape_nom,
                    defaults={"ordre": ordre_val, "montant_prevu": montant_d}
                )

                Versement.objects.create(
                    projet=projet, phase=phase,
                    montant=montant_d, date_versement=date_v,
                    type_versement=normalize_moyen_versement(moyen),
                    libelle=designation_txt,
                    reference_paiement=f"V-{basename}-{nb_versements+1:03d}",
                )
                nb_versements += 1

            self.stdout.write(self.style.SUCCESS(f"  + {nb_versements} versements"))

        # ── 7. COMMANDE MATERIAU => materiaux crees ─────────────────────
        nb_commandes = 0
        sheet_commande = None
        for name in wb.sheetnames:
            if "commande" in name.lower() and "mat" in name.lower():
                sheet_commande = name
                break

        if sheet_commande:
            ws = wb[sheet_commande]
            for row in ws.iter_rows(min_row=4, max_row=ws.max_row, max_col=8, values_only=True):
                etape, designation, date_val, unite, qte, total, _, cout = row
                if not designation or not qte:
                    continue
                date_c = to_date(date_val)
                if not date_c:
                    continue
                qte_d = to_decimal(qte)
                if qte_d <= 0:
                    continue

                cat_nom, mat_nom, unite_default = normalize_materiau_commande(designation)
                unite_str = normalize(unite) or unite_default
                get_or_create_materiel(cat_nom, mat_nom, unite_str)
                nb_commandes += 1

            self.stdout.write(self.style.SUCCESS(
                f"  + {nb_commandes} commandes materiaux analysees"
            ))

        # ── 8. ETAT STOCK => BonSortie ──────────────────────────────────
        nb_sorties = 0
        nb_bons = 0
        if "Etat stock" in wb.sheetnames:
            ws = wb["Etat stock"]
            current_bon = None
            current_date = None

            for row in ws.iter_rows(min_row=4, max_row=ws.max_row, max_col=13, values_only=True):
                if len(row) < 8:
                    continue
                _, etape, designation, date_val, unite, qte, facture, total_util, *rest = row

                if not designation or not qte:
                    continue
                date_s = to_date(date_val)
                if not date_s:
                    continue
                qte_d = to_decimal(qte)
                if qte_d <= 0:
                    continue

                cat_nom, mat_nom, unite_default = normalize_materiau_commande(designation)
                unite_str = normalize(unite) or unite_default
                materiel = get_or_create_materiel(cat_nom, mat_nom, unite_str)

                etape_str = normalize(etape) or "Travaux"

                # Un bon par jour
                if date_s != current_date:
                    current_bon = BonSortie(
                        projet=projet,
                        date_sortie=date_s,
                        responsable="Chef de chantier",
                        observations=f"Utilisation materiaux - {etape_str}",
                    )
                    current_bon.save()
                    current_date = date_s
                    nb_bons += 1

                facture_ref = normalize(facture) or ""
                LigneBonSortie.objects.create(
                    bon=current_bon,
                    materiel=materiel,
                    quantite=qte_d,
                    commentaire=f"{etape_str} - {mat_nom}" + (f" ({facture_ref})" if facture_ref else ""),
                )
                nb_sorties += 1

            self.stdout.write(self.style.SUCCESS(
                f"  + {nb_sorties} sorties stock ({nb_bons} bons de sortie)"
            ))

        # ── 9. DEPENSES ─────────────────────────────────────────────────
        nb_achats = 0

        if "DEPENSES" in wb.sheetnames:
            ws = wb["DEPENSES"]
            for row in ws.iter_rows(min_row=4, max_row=ws.max_row, max_col=13, values_only=True):
                no, etape, lot, designation, intervenant, moyen, date_val, montant, frais, facture, depense_total, dep_cumul, comment = row

                date_a = fix_date(to_date(date_val))
                montant_d = to_decimal(montant)
                if not date_a or montant_d <= 0:
                    continue

                designation_txt = normalize(designation) or "Depense"
                lot_norm = normalize_lot(lot)
                etape_nom = normalize(etape) or "Travaux"
                moyen_norm = normalize_moyen_achat(moyen)
                frais_d = to_decimal(frais)
                total_depense = to_decimal(depense_total) if depense_total else montant_d + frais_d

                # Classifier l'intervenant
                nom_affiche, role = classify_intervenant(intervenant)

                # Extraire quantite
                qty, unite_ext, _ = extract_quantity_and_details(designation, lot_norm)

                # ── Creer l'intervenant dans la BONNE table ──────────────
                fournisseur_obj = None

                if role == "macon":
                    get_or_create_employe(nom_affiche)
                    # Pas de fournisseur pour un macon => fournisseur=None

                elif role == "sous_traitant":
                    st = get_or_create_sous_traitant(nom_affiche)
                    # Cumuler pour contrat
                    if nom_affiche not in contrats_st_data:
                        contrats_st_data[nom_affiche] = {
                            "st": st, "montant": Decimal("0"),
                            "lots": set(), "date_debut": date_a,
                        }
                    contrats_st_data[nom_affiche]["montant"] += total_depense
                    contrats_st_data[nom_affiche]["lots"].add(lot_norm)
                    if date_a < contrats_st_data[nom_affiche]["date_debut"]:
                        contrats_st_data[nom_affiche]["date_debut"] = date_a
                    # Pas de fournisseur pour un sous-traitant => fournisseur=None

                elif role == "frais":
                    # Frais generaux => pas de fournisseur
                    pass

                else:
                    # Fournisseur classique
                    fournisseur_obj = get_or_create_fournisseur(nom_affiche)

                # ── Creer l'achat ────────────────────────────────────────
                facture_ref = normalize(facture) or f"D-{basename}-{nb_achats+1:04d}"

                achat = Achat(
                    date_achat=date_a,
                    projet=projet,
                    fournisseur=fournisseur_obj,
                    mode_paiement=moyen_norm,
                    numero_facture=facture_ref,
                    tva_active=False,
                )
                achat.save()

                # Materiau
                materiel = materiaux_cache.get(lot_norm)
                if not materiel:
                    cat = categories_cache.get(lot_norm)
                    if not cat:
                        cat, _ = CategorieMateriel.objects.get_or_create(nom=lot_norm)
                        categories_cache[lot_norm] = cat
                    materiel, _ = Materiel.objects.get_or_create(
                        nom=lot_norm, defaults={"unite": "forfait", "categorie": cat}
                    )
                    materiaux_cache[lot_norm] = materiel

                # Prix unitaire
                prix_unitaire = total_depense / Decimal(str(qty)) if qty > 1 else total_depense

                LigneAchat.objects.create(
                    achat=achat,
                    materiel=materiel,
                    quantite=int(qty),
                    prix_unitaire=prix_unitaire,
                    commentaire=f"{etape_nom} - {designation_txt}",
                )

                achat.calcul_totaux()
                achat.save(update_fields=["total_ht", "total_tva", "total_ttc"])
                nb_achats += 1

            self.stdout.write(self.style.SUCCESS(
                f"  + {nb_achats} depenses importees"
            ))

        # ── 10. Contrats sous-traitance ─────────────────────────────────
        nb_contrats = 0
        for nom_st, info in contrats_st_data.items():
            lot_desc = " / ".join(sorted(info["lots"]))
            contrat, created = ContratSousTraitance.objects.get_or_create(
                projet=projet, sous_traitant=info["st"],
                defaults={
                    "lot": lot_desc,
                    "montant": info["montant"],
                    "montant_paye": info["montant"],
                    "date_debut": info["date_debut"],
                    "statut": "en_cours",
                }
            )
            if created:
                nb_contrats += 1
                self.stdout.write(
                    f"  + Contrat : {nom_st} => {lot_desc} ({info['montant']:,.0f} FCFA)"
                )

        # ── 11. Plan comptable SYSCOHADA ────────────────────────────────
        plan_comptable = [
            # Classe 1 - Ressources durables
            ("101000", "Capital social", "passif", 1),
            ("164000", "Emprunts aupres des etablissements de credit", "passif", 1),
            # Classe 2 - Actif immobilise
            ("213000", "Constructions", "actif", 2),
            ("215000", "Materiel et outillage", "actif", 2),
            ("231000", "Batiments en cours", "actif", 2),
            # Classe 3 - Stocks
            ("311000", "Matieres premieres", "actif", 3),
            ("321000", "Matieres consommables", "actif", 3),
            ("371000", "Stocks de marchandises", "actif", 3),
            # Classe 4 - Tiers
            ("401000", "Fournisseurs", "passif", 4),
            ("408000", "Fournisseurs - factures non parvenues", "passif", 4),
            ("411000", "Clients", "actif", 4),
            ("421000", "Personnel - remunerations dues", "passif", 4),
            ("431000", "Organismes sociaux", "passif", 4),
            ("445000", "TVA deductible", "actif", 4),
            ("447000", "TVA collectee", "passif", 4),
            # Classe 5 - Tresorerie
            ("521000", "Banque", "actif", 5),
            ("571000", "Caisse", "actif", 5),
            ("585000", "Mobile money (Wave/OM)", "actif", 5),
            # Classe 6 - Charges
            ("601000", "Achats matieres premieres et fournitures", "charge", 6),
            ("604000", "Achats de materiaux et equipements", "charge", 6),
            ("605000", "Autres achats", "charge", 6),
            ("611000", "Sous-traitance generale", "charge", 6),
            ("613000", "Locations et charges locatives", "charge", 6),
            ("616000", "Transports de biens", "charge", 6),
            ("621000", "Personnel interimaire et main d'oeuvre", "charge", 6),
            ("641000", "Charges de personnel", "charge", 6),
            ("661000", "Charges d'interets", "charge", 6),
            ("681000", "Dotations aux amortissements", "charge", 6),
            # Classe 7 - Produits
            ("706000", "Travaux", "produit", 7),
            ("707000", "Ventes de marchandises", "produit", 7),
            ("758000", "Autres produits d'exploitation", "produit", 7),
        ]

        nb_comptes = 0
        for code, libelle, type_c, classe in plan_comptable:
            _, created = CompteComptable.objects.get_or_create(
                code=code,
                defaults={"libelle": libelle, "type_compte": type_c, "classe": classe}
            )
            if created:
                nb_comptes += 1

        if nb_comptes:
            self.stdout.write(self.style.SUCCESS(
                f"  + {nb_comptes} comptes SYSCOHADA crees"
            ))

        # ── 12. Compte bancaire et caisse ───────────────────────────────
        compte_banque, created = CompteBancaire.objects.get_or_create(
            nom="Banque principale",
            defaults={
                "type_compte": "banque",
                "banque": "Banque du projet",
                "numero_compte": "SN001-LAMANE",
                "solde_initial": Decimal("0"),
                "actif": True,
            }
        )
        if created:
            self.stdout.write(f"  + Compte bancaire : Banque principale")

        compte_caisse, created = CompteBancaire.objects.get_or_create(
            nom="Caisse chantier",
            defaults={
                "type_compte": "caisse",
                "banque": "",
                "numero_compte": "CAISSE-001",
                "solde_initial": Decimal("0"),
                "actif": True,
            }
        )
        if created:
            self.stdout.write(f"  + Compte caisse : Caisse chantier")

        compte_mobile, created = CompteBancaire.objects.get_or_create(
            nom="Wave / Mobile money",
            defaults={
                "type_compte": "mobile_money",
                "banque": "Wave",
                "numero_compte": "WAVE-001",
                "solde_initial": Decimal("0"),
                "actif": True,
            }
        )
        if created:
            self.stdout.write(f"  + Compte mobile : Wave / Mobile money")

        # ── 13. Ecritures comptables ────────────────────────────────────
        nb_ecritures = 0

        # Ecritures pour les versements
        for v in Versement.objects.filter(projet=projet):
            ec = generer_ecriture_versement(v)
            if ec:
                nb_ecritures += 1
                # Transaction bancaire correspondante
                if v.type_versement in ("virement bancaire", "cheque"):
                    cpt = compte_banque
                elif v.type_versement == "wave":
                    cpt = compte_mobile
                else:
                    cpt = compte_caisse
                TransactionBancaire.objects.get_or_create(
                    versement=v,
                    defaults={
                        "compte": cpt,
                        "type_transaction": "entree",
                        "montant": v.montant,
                        "date_transaction": v.date_versement,
                        "libelle": f"Versement client - {v.libelle}",
                        "projet": projet,
                    }
                )

        # Ecritures pour les achats
        for a in Achat.objects.filter(projet=projet):
            ec = generer_ecriture_achat(a)
            if ec:
                nb_ecritures += 1
                # Transaction bancaire (sortie)
                if a.mode_paiement == "virement":
                    cpt = compte_banque
                elif a.mode_paiement == "autre":
                    cpt = compte_mobile
                else:
                    cpt = compte_caisse
                TransactionBancaire.objects.get_or_create(
                    achat=a,
                    defaults={
                        "compte": cpt,
                        "type_transaction": "sortie",
                        "montant": a.total_ttc or Decimal("0"),
                        "date_transaction": a.date_achat,
                        "libelle": f"Achat - {a.numero_facture}",
                        "projet": projet,
                    }
                )

        # Ecritures pour les contrats sous-traitance
        for c in ContratSousTraitance.objects.filter(projet=projet):
            ec = generer_ecriture_sous_traitance(c)
            if ec:
                nb_ecritures += 1

        self.stdout.write(self.style.SUCCESS(
            f"  + {nb_ecritures} ecritures comptables generees"
        ))
        self.stdout.write(self.style.SUCCESS(
            f"  + {TransactionBancaire.objects.filter(projet=projet).count()} transactions bancaires"
        ))

        # ── Resume ──────────────────────────────────────────────────────
        total_versements = sum(v.montant for v in Versement.objects.filter(projet=projet))
        total_depenses = sum(a.total_ttc for a in Achat.objects.filter(projet=projet) if a.total_ttc)

        self.stdout.write(f"\n{'='*60}")
        self.stdout.write(f"RESUME IMPORT - {nom_projet}")
        self.stdout.write(f"{'='*60}")
        self.stdout.write(f"  Proprietaire    : {client}")
        self.stdout.write(f"  Versements      : {nb_versements} => {total_versements:,.0f} FCFA")
        self.stdout.write(f"  Depenses        : {nb_achats} => {total_depenses:,.0f} FCFA")
        self.stdout.write(f"  Solde           : {total_versements - total_depenses:,.0f} FCFA")
        self.stdout.write(f"  Employes/Macons : {len(employes_cache)}")
        self.stdout.write(f"  Sous-traitants  : {len(sous_traitants_cache)} ({nb_contrats} contrats)")
        self.stdout.write(f"  Fournisseurs    : {len(fournisseurs_cache)}")
        self.stdout.write(f"  Bons de sortie  : {nb_bons} ({nb_sorties} lignes)")
        self.stdout.write(f"  Ecritures compt.: {EcritureComptable.objects.count()}")
        self.stdout.write(f"  Transactions    : {TransactionBancaire.objects.filter(projet=projet).count()}")
        self.stdout.write(f"  Plan comptable  : {CompteComptable.objects.count()} comptes")
        self.stdout.write(f"  Categories      : {CategorieMateriel.objects.count()}")
        self.stdout.write(f"  Materiaux       : {Materiel.objects.count()}")
        self.stdout.write(f"{'='*60}\n")
        self.stdout.write(self.style.SUCCESS("Import termine avec succes !"))
